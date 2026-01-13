from fastapi import APIRouter, Depends, HTTPException, Request, Form, UploadFile, File
from typing import Optional, List
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from firebase_admin import firestore
from datetime import datetime
from io import BytesIO
from app.database import get_db
from app import schemas
from app.jinja_templates import templates
from app.firestore_models import document_to_dict, datetime_to_timestamp
from app.storage import storage_client
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()
html_router = APIRouter()


@router.get("/", response_model=list[schemas.Product])
def get_products(skip: int = 0, limit: int = 100, db = Depends(get_db)):
    products_ref = db.collection("products")
    docs = products_ref.order_by("created_at", direction=firestore.Query.DESCENDING).offset(skip).limit(limit).stream()
    products = []
    for doc in docs:
        product = document_to_dict(doc)
        if product:
            products.append(product)
    return products


@router.get("/{product_id}", response_model=schemas.Product)
def get_product(product_id: str, db = Depends(get_db)):
    doc = db.collection("products").document(product_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    return document_to_dict(doc)


@router.post("/", response_model=schemas.Product)
def create_product(product: schemas.ProductCreate, db = Depends(get_db)):
    doc_ref = db.collection("products").document()
    data = product.model_dump()
    # Convert ProductType enum to string value for Firestore
    if "product_type" in data and data["product_type"] is not None:
        if hasattr(data["product_type"], "value"):
            data["product_type"] = data["product_type"].value
    data["created_at"] = firestore.SERVER_TIMESTAMP
    doc_ref.set(data)
    doc = doc_ref.get()
    return document_to_dict(doc)


@router.put("/{product_id}", response_model=schemas.Product)
def update_product(product_id: str, product: schemas.ProductUpdate, db = Depends(get_db)):
    doc_ref = db.collection("products").document(product_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product.model_dump(exclude_unset=True)
    # Convert ProductType enum to string value for Firestore
    if "product_type" in update_data and update_data["product_type"] is not None:
        if hasattr(update_data["product_type"], "value"):
            update_data["product_type"] = update_data["product_type"].value
    if update_data:
        doc_ref.update(update_data)
    
    updated_doc = doc_ref.get()
    return document_to_dict(updated_doc)


@router.delete("/{product_id}")
def delete_product(product_id: str, db = Depends(get_db)):
    doc_ref = db.collection("products").document(product_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Delete related BOM lines
    bom_lines = db.collection("product_bom").where("product_id", "==", product_id).stream()
    for bom_line in bom_lines:
        bom_line.reference.delete()
    
    # Delete related images
    images = db.collection("product_images").where("product_id", "==", product_id).stream()
    for image in images:
        image_data = image.to_dict()
        if image_data and "image_url" in image_data:
            # Delete from storage if using GCS
            storage_client.delete_file(image_data["image_url"])
        image.reference.delete()
    
    # Delete product
    doc_ref.delete()
    return {"message": "Product deleted"}


# BOM endpoints
@router.get("/{product_id}/bom", response_model=list[schemas.ProductBOM])
def get_product_bom(product_id: str, db = Depends(get_db)):
    bom_ref = db.collection("product_bom")
    docs = bom_ref.where("product_id", "==", product_id).stream()
    bom_lines = []
    for doc in docs:
        bom = document_to_dict(doc)
        if bom:
            bom_lines.append(bom)
    return bom_lines


@router.post("/{product_id}/bom", response_model=schemas.ProductBOM)
def create_bom_line(product_id: str, bom: schemas.ProductBOMCreate, db = Depends(get_db)):
    # Verify product exists
    product_doc = db.collection("products").document(product_id).get()
    if not product_doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Verify material exists
    material_doc = db.collection("materials").document(bom.material_id).get()
    if not material_doc.exists:
        raise HTTPException(status_code=404, detail="Material not found")
    
    # Verify purchase exists and belongs to the material
    purchase_doc = db.collection("purchases").document(bom.purchase_id).get()
    if not purchase_doc.exists:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    purchase_data = purchase_doc.to_dict()
    if purchase_data and purchase_data.get("material_id") != bom.material_id:
        raise HTTPException(status_code=400, detail="Purchase does not match material")
    
    # Get unit from material
    material_data = material_doc.to_dict()
    unit = material_data.get("unit") if material_data else bom.unit
    
    doc_ref = db.collection("product_bom").document()
    data = bom.model_dump()
    data["product_id"] = product_id
    data["unit"] = unit
    doc_ref.set(data)
    doc = doc_ref.get()
    return document_to_dict(doc)


@router.delete("/bom/{bom_id}")
def delete_bom_line(bom_id: str, db = Depends(get_db)):
    doc_ref = db.collection("product_bom").document(bom_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="BOM line not found")
    
    doc_ref.delete()
    return {"message": "BOM line deleted"}


# HTML routes
@html_router.get("/products", response_class=HTMLResponse)
async def products_page(request: Request, page: int = 1, per_page: int = 10, db = Depends(get_db)):
    # Calculate skip
    skip = (page - 1) * per_page
    
    # Get total count
    products_ref = db.collection("products")
    total_docs = products_ref.stream()
    total_count = sum(1 for _ in total_docs)
    
    # Get paginated products
    docs = products_ref.order_by("created_at", direction=firestore.Query.DESCENDING).offset(skip).limit(per_page).stream()
    products = []
    for doc in docs:
        product = document_to_dict(doc)
        if product:
            products.append(product)
    
    # Calculate pagination info
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
    
    # Generate page numbers for pagination (show current page ± 2, plus first and last)
    page_numbers = []
    if total_pages <= 7:
        # Show all pages if 7 or fewer
        page_numbers = list(range(1, total_pages + 1))
    else:
        # Show first page
        page_numbers.append(1)
        # Show pages around current
        start = max(2, page - 2)
        end = min(total_pages - 1, page + 2)
        if start > 2:
            page_numbers.append(None)  # Ellipsis marker
        page_numbers.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_numbers.append(None)  # Ellipsis marker
        # Show last page
        if total_pages > 1:
            page_numbers.append(total_pages)
    
    return templates.TemplateResponse("products.html", {
        "request": request,
        "products": products,
        "page": page,
        "per_page": per_page,
        "total_count": total_count,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "page_numbers": page_numbers
    })


@html_router.get("/products/new", response_class=HTMLResponse)
async def new_product_page(request: Request):
    return templates.TemplateResponse("product_form.html", {"request": request, "product": None})


# Define delete-all route BEFORE {product_id} routes to ensure proper matching
@html_router.post("/products/delete-all", response_class=HTMLResponse)
async def delete_all_products_form(request: Request, db = Depends(get_db)):
    """Delete all products and their related data (BOM lines, images)"""
    try:
        products_ref = db.collection("products")
        products = products_ref.stream()
        
        deleted_count = 0
        errors = []
        
        # Collect all product IDs first
        product_ids = []
        for product_doc in products:
            if product_doc.exists:
                product_ids.append(product_doc.id)
        
        # If no products, just redirect
        if not product_ids:
            return RedirectResponse(url="/products", status_code=303)
        
        # Delete each product and its related data
        for product_id in product_ids:
            try:
                # Check if product exists
                doc_ref = db.collection("products").document(product_id)
                doc = doc_ref.get()
                if not doc.exists:
                    continue  # Skip if already deleted
                
                # Delete related BOM lines
                bom_lines = db.collection("product_bom").where("product_id", "==", product_id).stream()
                for bom_line in bom_lines:
                    try:
                        bom_line.reference.delete()
                    except Exception as bom_error:
                        print(f"Warning: Could not delete BOM line {bom_line.id}: {bom_error}")
                
                # Delete related images
                images = db.collection("product_images").where("product_id", "==", product_id).stream()
                for image in images:
                    try:
                        image_data = image.to_dict()
                        if image_data and "image_url" in image_data:
                            # Delete from storage if using GCS
                            try:
                                storage_client.delete_file(image_data["image_url"])
                            except Exception as storage_error:
                                print(f"Warning: Could not delete image from storage: {storage_error}")
                        image.reference.delete()
                    except Exception as image_error:
                        print(f"Warning: Could not delete image {image.id}: {image_error}")
                
                # Delete product
                doc_ref.delete()
                deleted_count += 1
                
            except Exception as e:
                error_msg = str(e)
                # Don't treat "not found" as an error - product might have been deleted already
                if "not found" not in error_msg.lower():
                    errors.append(f"Product {product_id}: {error_msg}")
                    print(f"Error deleting product {product_id}: {e}")
                # Continue with other products even if one fails
        
        if errors:
            print(f"Errors during bulk delete: {errors}")
        
        print(f"Successfully deleted {deleted_count} product(s)")
        
    except Exception as e:
        print(f"Critical error in delete_all_products_form: {e}")
        import traceback
        traceback.print_exc()
        # Still redirect even if there's an error
    
    return RedirectResponse(url="/products", status_code=303)


@html_router.get("/products/{product_id}", response_class=HTMLResponse)
async def product_detail_page(request: Request, product_id: str, db = Depends(get_db)):
    doc = db.collection("products").document(product_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    product = document_to_dict(doc)
    
    # Get BOM lines
    bom_ref = db.collection("product_bom")
    bom_docs = bom_ref.where("product_id", "==", product_id).stream()
    bom_lines = []
    for bom_doc in bom_docs:
        bom = document_to_dict(bom_doc)
        if bom:
            # Get material name
            material_doc = db.collection("materials").document(bom["material_id"]).get()
            if material_doc.exists:
                material = document_to_dict(material_doc)
                bom["material_name"] = material["name"] if material else "Unknown"
            else:
                bom["material_name"] = "Unknown"
            
            # Get purchase info
            purchase_doc = db.collection("purchases").document(bom["purchase_id"]).get()
            if purchase_doc.exists:
                purchase = document_to_dict(purchase_doc)
                if purchase:
                    purchase_date = purchase.get("purchase_date")
                    if purchase_date:
                        if hasattr(purchase_date, 'strftime'):
                            date_str = purchase_date.strftime('%Y-%m-%d')
                        else:
                            date_str = str(purchase_date)
                    else:
                        date_str = "Unknown"
                    bom["purchase_info"] = f"{purchase['supplier_name']} - {purchase['unit_cost']} {purchase['currency']} ({date_str})"
                else:
                    bom["purchase_info"] = "Unknown"
            else:
                bom["purchase_info"] = "Unknown"
            bom_lines.append(bom)
    
    # Get materials for dropdown
    materials_ref = db.collection("materials")
    mat_docs = materials_ref.order_by("name").stream()
    materials = []
    for mat_doc in mat_docs:
        material = document_to_dict(mat_doc)
        if material:
            materials.append(material)
    
    # Get all purchases for the materials dropdown
    purchases_ref = db.collection("purchases")
    purchase_docs = purchases_ref.order_by("purchase_date", direction=firestore.Query.DESCENDING).stream()
    purchases = []
    for purchase_doc in purchase_docs:
        purchase = document_to_dict(purchase_doc)
        if purchase:
            purchases.append(purchase)
    
    # Get product images - fetch without order_by to avoid index requirement, then sort in Python
    images_ref = db.collection("product_images")
    image_docs = images_ref.where("product_id", "==", product_id).stream()
    product_images = []
    print(f"Debug: Fetching images for product_id: {product_id}")
    image_count = 0
    for image_doc in image_docs:
        image = document_to_dict(image_doc)
        if image:
            image_count += 1
            # Generate signed URL for the image if needed
            original_url = image.get('image_url', '')
            if original_url:
                # Try to get signed URL for better access
                signed_url = storage_client.get_signed_url(original_url)
                image['image_url'] = signed_url
            print(f"Debug: Found image {image_count}: id={image.get('id')}, url={image.get('image_url', 'N/A')[:80]}..., order={image.get('order', 'N/A')}")
            product_images.append(image)
    print(f"Debug: Total images found: {len(product_images)}")
    # Sort by order, then by created_at if order is the same (sorting in Python)
    def sort_key(x):
        order = x.get("order", 0)
        created_at = x.get("created_at")
        # Handle Firestore timestamp or datetime
        if created_at:
            if hasattr(created_at, 'timestamp'):
                created_at_val = created_at.timestamp()
            elif isinstance(created_at, datetime):
                created_at_val = created_at.timestamp()
            else:
                created_at_val = 0
        else:
            created_at_val = 0
        return (order, created_at_val)
    product_images.sort(key=sort_key)
    
    return templates.TemplateResponse(
        "product_detail.html",
        {
            "request": request,
            "product": product,
            "bom_lines": bom_lines,
            "materials": materials,
            "purchases": purchases,
            "product_images": product_images
        }
    )


@html_router.get("/products/{product_id}/edit", response_class=HTMLResponse)
async def edit_product_page(request: Request, product_id: str, db = Depends(get_db)):
    doc = db.collection("products").document(product_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    product = document_to_dict(doc)
    return templates.TemplateResponse("product_form.html", {"request": request, "product": product})


@html_router.post("/products", response_class=HTMLResponse)
async def create_product_form(
    request: Request,
    db = Depends(get_db)
):
    # Parse form data manually to handle multiple files
    # Important: Use await request.form() to get multipart form data
    form = await request.form()
    print(f"Debug: Form keys: {list(form.keys())}")
    
    sku = form.get("sku")
    name = form.get("name")
    description = form.get("description")
    count = int(form.get("count", 1))
    collection_name = form.get("collection_name")
    product_type = form.get("product_type")
    
    if not sku or not name:
        raise HTTPException(status_code=400, detail="SKU and name are required")
    
    # Convert product_type string to enum if provided
    product_type_enum = None
    if product_type and product_type.strip():
        try:
            from app.models import ProductType
            product_type_enum = ProductType(product_type)
        except ValueError:
            # Invalid product type, ignore it
            product_type_enum = None
    
    product = schemas.ProductCreate(
        sku=sku,
        name=name,
        description=description,
        image_url=None,  # No longer using image_url
        count=count,
        collection_name=collection_name if collection_name and collection_name.strip() else None,
        product_type=product_type_enum
    )
    result = create_product(product, db)
    product_id = result['id']
    
    # Handle multiple image uploads
    # FastAPI/Starlette: when multiple files have same name, they come as separate entries in multi_items()
    image_files = []
    try:
        # Method 1: Use multi_items() - this should return ALL entries including duplicates
        form_items = list(form.multi_items())
        print(f"Debug: Total form items: {len(form_items)}")
        
        # Count how many "images" entries we have
        images_count = sum(1 for key, _ in form_items if key == "images")
        print(f"Debug: Found {images_count} 'images' entries in form")
        
        # Collect all "images" entries - they should all be separate entries
        for key, value in form_items:
            if key == "images":
                print(f"Debug: Found 'images' entry: type={type(value).__name__}, filename={getattr(value, 'filename', 'N/A')}")
                # Check if it's an UploadFile instance
                if isinstance(value, UploadFile):
                    if value.filename and value.filename.strip():
                        # IMPORTANT: Create a copy or store the file content immediately
                        # because UploadFile can only be read once
                        image_files.append(value)
                        print(f"Debug: ✓ Added file: {value.filename}")
                # Also check for file-like objects
                elif hasattr(value, 'filename') and value.filename and value.filename.strip():
                    image_files.append(value)
                    print(f"Debug: ✓ Added file (hasattr): {value.filename}")
        
        # Method 2: If still only one file, try accessing form's internal structure
        # Sometimes multiple files with same name get grouped
        if len(image_files) == 1 and images_count > 1:
            print(f"Debug: Only 1 file collected but {images_count} entries found - trying alternative method")
            try:
                # Try to access the raw multipart data
                if hasattr(form, '_list'):
                    print(f"Debug: form._list has {len(form._list)} items")
                    for idx, item in enumerate(form._list):
                        print(f"Debug: _list[{idx}]: name={getattr(item, 'name', 'N/A')}, type={type(item).__name__}")
                        if hasattr(item, 'name') and item.name == "images":
                            if hasattr(item, 'file') and item.file:
                                if hasattr(item.file, 'filename') and item.file.filename:
                                    if item.file not in image_files:
                                        image_files.append(item.file)
                                        print(f"Debug: ✓ Added file via _list: {item.file.filename}")
            except Exception as e:
                print(f"Debug: Could not access form._list: {e}")
        
        print(f"Debug: ✓ Total image files collected: {len(image_files)}")
        if image_files:
            print(f"Debug: Files: {[f.filename for f in image_files]}")
    except Exception as e:
        print(f"✗ Warning: Error parsing image files from form: {e}")
        import traceback
        traceback.print_exc()
    
    # Upload images if provided
    if image_files:
        print(f"Uploading {len(image_files)} image(s) for product {product_id} (SKU: {sku})")
        uploaded_files = []
        try:
            images_ref = db.collection("product_images")
            image_docs = images_ref.where("product_id", "==", product_id).stream()
            max_order = -1
            for image_doc in image_docs:
                image_data = image_doc.to_dict()
                if image_data and "order" in image_data:
                    max_order = max(max_order, image_data["order"])
            
            for idx, image_file in enumerate(image_files):
                try:
                    if not image_file.filename or not image_file.filename.strip():
                        print(f"⚠ Warning: Skipping file with empty filename at index {idx}")
                        continue
                    
                    print(f"Debug: Processing image {idx + 1}/{len(image_files)}: {image_file.filename}")
                    
                    # Check if storage is available
                    if not storage_client.bucket:
                        print(f"⚠ Warning: Firebase Storage not configured. Skipping image upload for {image_file.filename}")
                        print(f"Hint: Set FIREBASE_STORAGE_BUCKET in your .env file")
                        continue
                    
                    # Read file content - IMPORTANT: Reset file pointer if needed
                    # FastAPI UploadFile can only be read once, so we need to read it here
                    try:
                        # Reset to beginning if possible
                        if hasattr(image_file.file, 'seek'):
                            image_file.file.seek(0)
                        file_content = await image_file.read()
                        print(f"Debug: Read {len(file_content)} bytes from {image_file.filename}")
                    except Exception as read_error:
                        print(f"✗ Error reading file {image_file.filename}: {read_error}")
                        continue
                    
                    # Determine content type
                    content_type = image_file.content_type or "image/jpeg"
                    print(f"Debug: Content type: {content_type}")
                    
                    # Upload to Firebase Storage - organize by SKU
                    image_url = storage_client.upload_file(
                        file_content, 
                        image_file.filename, 
                        content_type,
                        sku=sku  # Pass SKU to organize files in subfolder
                    )
                    
                    if image_url:
                        # Create product image record
                        doc_ref = db.collection("product_images").document()
                        data = {
                            "product_id": product_id,
                            "image_url": image_url,
                            "order": max_order + 1 + idx,
                            "created_at": firestore.SERVER_TIMESTAMP
                        }
                        doc_ref.set(data)
                        uploaded_files.append(image_file.filename)
                        print(f"✓ Successfully uploaded and saved image {idx + 1}: {image_file.filename} -> {image_url}")
                    else:
                        print(f"⚠ Warning: upload_file returned None for {image_file.filename}")
                except Exception as e:
                    print(f"✗ Error uploading image {image_file.filename}: {e}")
                    import traceback
                    traceback.print_exc()
                    # Continue with other images even if one fails
                    continue
            
            if uploaded_files:
                print(f"✓ Successfully uploaded {len(uploaded_files)} image(s): {', '.join(uploaded_files)}")
        except Exception as e:
            print(f"✗ Error processing images: {e}")
            import traceback
            traceback.print_exc()
            # Don't fail the entire product creation if images fail
    
    return RedirectResponse(url=f"/products/{product_id}", status_code=303)


@html_router.post("/products/{product_id}", response_class=HTMLResponse)
async def update_product_form(
    request: Request,
    product_id: str,
    sku: str = Form(None),
    name: str = Form(None),
    description: str = Form(None),
    image_url: str = Form(None),
    count: int = Form(None),
    collection_name: str = Form(None),
    product_type: str = Form(None),
    db = Depends(get_db)
):
    update_data = {}
    if sku:
        update_data["sku"] = sku
    if name:
        update_data["name"] = name
    if description is not None:
        update_data["description"] = description
    if image_url is not None:
        update_data["image_url"] = image_url
    if count is not None:
        update_data["count"] = count
    if collection_name is not None:
        update_data["collection_name"] = collection_name if collection_name.strip() else None
    if product_type is not None:
        # Convert product_type string to enum if provided
        if product_type and product_type.strip():
            try:
                from app.models import ProductType
                update_data["product_type"] = ProductType(product_type)
            except ValueError:
                # Invalid product type, set to None
                update_data["product_type"] = None
        else:
            update_data["product_type"] = None
    
    product = schemas.ProductUpdate(**update_data)
    update_product(product_id, product, db)
    return RedirectResponse(url=f"/products/{product_id}", status_code=303)


@html_router.post("/products/{product_id}/bom", response_class=HTMLResponse)
async def create_bom_line_form(
    request: Request,
    product_id: str,
    material_id: str = Form(...),
    purchase_id: str = Form(...),
    qty_required: float = Form(...),
    note: str = Form(None),
    db = Depends(get_db)
):
    # Get unit from material
    material_doc = db.collection("materials").document(material_id).get()
    if not material_doc.exists:
        raise HTTPException(status_code=404, detail="Material not found")
    
    material_data = material_doc.to_dict()
    unit = material_data.get("unit") if material_data else ""
    
    bom = schemas.ProductBOMCreate(
        product_id=product_id,
        material_id=material_id,
        purchase_id=purchase_id,
        qty_required=qty_required,
        unit=unit,
        note=note
    )
    create_bom_line(product_id, bom, db)
    return RedirectResponse(url=f"/products/{product_id}", status_code=303)


@router.post("/bulk-update-counts")
def bulk_update_product_counts(request: schemas.BulkCountUpdateRequest, db = Depends(get_db)):
    """
    Bulk update product counts.
    """
    updated_count = 0
    for update_item in request.updates:
        product_id = update_item.product_id
        count = update_item.count
        
        doc_ref = db.collection("products").document(product_id)
        doc = doc_ref.get()
        if doc.exists:
            doc_ref.update({"count": count})
            updated_count += 1
        else:
            raise HTTPException(status_code=404, detail=f"Product {product_id} not found")
    
    return {"message": f"Updated {updated_count} product(s)", "updated": updated_count}


@html_router.post("/products/{product_id}/update-count", response_class=HTMLResponse)
async def update_product_count_form(
    request: Request,
    product_id: str,
    count: int = Form(...),
    db = Depends(get_db)
):
    update_data = {"count": count}
    product = schemas.ProductUpdate(**update_data)
    update_product(product_id, product, db)
    return RedirectResponse(url="/products", status_code=303)


@html_router.post("/products/{product_id}/delete", response_class=HTMLResponse)
async def delete_product_form(request: Request, product_id: str, db = Depends(get_db)):
    delete_product(product_id, db)
    return RedirectResponse(url="/products", status_code=303)


@router.post("/{product_id}/images", response_model=schemas.ProductImage)
async def add_product_image(
    product_id: str,
    image_url: str = Form(...),
    db = Depends(get_db)
):
    """Add an image to a product"""
    product_doc = db.collection("products").document(product_id).get()
    if not product_doc.exists:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get the highest order number for this product
    images_ref = db.collection("product_images")
    image_docs = images_ref.where("product_id", "==", product_id).stream()
    max_order = -1
    for image_doc in image_docs:
        image_data = image_doc.to_dict()
        if image_data and "order" in image_data:
            max_order = max(max_order, image_data["order"])
    
    doc_ref = db.collection("product_images").document()
    data = {
        "product_id": product_id,
        "image_url": image_url,
        "order": max_order + 1,
        "created_at": firestore.SERVER_TIMESTAMP
    }
    doc_ref.set(data)
    doc = doc_ref.get()
    return document_to_dict(doc)


@router.delete("/images/{image_id}")
def delete_product_image(image_id: str, db = Depends(get_db)):
    """Delete a product image"""
    doc_ref = db.collection("product_images").document(image_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Image not found")
    
    image_data = doc.to_dict()
    if image_data and "image_url" in image_data:
        # Delete from storage if using GCS
        storage_client.delete_file(image_data["image_url"])
    
    doc_ref.delete()
    return {"message": "Image deleted"}


@html_router.post("/products/{product_id}/images", response_class=HTMLResponse)
async def add_product_image_form(
    request: Request,
    product_id: str,
    image_url: str = Form(...),
    db = Depends(get_db)
):
    add_product_image(product_id, image_url, db)
    return RedirectResponse(url=f"/products/{product_id}", status_code=303)


@html_router.post("/products/images/{image_id}/delete", response_class=HTMLResponse)
async def delete_product_image_form(
    request: Request,
    image_id: str,
    db = Depends(get_db)
):
    image_doc = db.collection("product_images").document(image_id).get()
    if not image_doc.exists:
        raise HTTPException(status_code=404, detail="Image not found")
    
    image_data = image_doc.to_dict()
    product_id = image_data.get("product_id") if image_data else None
    delete_product_image(image_id, db)
    return RedirectResponse(url=f"/products/{product_id}", status_code=303)


@html_router.get("/products/export/excel")
async def export_products_excel(db = Depends(get_db)):
    """Export all products to Excel file"""
    products_ref = db.collection("products")
    docs = products_ref.order_by("created_at", direction=firestore.Query.DESCENDING).stream()
    products = []
    for doc in docs:
        product = document_to_dict(doc)
        if product:
            products.append(product)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Define header row
    headers = [
        "ID",
        "SKU",
        "Name",
        "Collection Name",
        "Product Type",
        "Description",
        "Count",
        "Image URL",
        "Created At",
        "BOM Materials Count"
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, product in enumerate(products, 2):
        # Get BOM count
        bom_ref = db.collection("product_bom")
        bom_docs = bom_ref.where("product_id", "==", product["id"]).stream()
        bom_count = sum(1 for _ in bom_docs)
        
        # Handle datetime
        created_at = product.get("created_at")
        if created_at:
            if hasattr(created_at, 'strftime'):
                created_at_str = created_at.strftime('%Y-%m-%d %H:%M:%S')
            else:
                created_at_str = str(created_at)
        else:
            created_at_str = ""
        
        # Format product type for display
        product_type = product.get("product_type", "")
        if product_type:
            if product_type == "RING":
                product_type_display = "Ring"
            elif product_type == "NECKLACE":
                product_type_display = "Necklace"
            elif product_type == "EARRING":
                product_type_display = "Earring"
            elif product_type == "BRACELET":
                product_type_display = "Bracelet"
            else:
                product_type_display = product_type
        else:
            product_type_display = ""
        
        ws.cell(row=row_num, column=1, value=product["id"])
        ws.cell(row=row_num, column=2, value=product["sku"])
        ws.cell(row=row_num, column=3, value=product["name"])
        ws.cell(row=row_num, column=4, value=product.get("collection_name", "") or "")
        ws.cell(row=row_num, column=5, value=product_type_display)
        ws.cell(row=row_num, column=6, value=product.get("description", "") or "")
        ws.cell(row=row_num, column=7, value=product.get("count", 1))
        ws.cell(row=row_num, column=8, value=product.get("image_url", "") or "")
        ws.cell(row=row_num, column=9, value=created_at_str)
        ws.cell(row=row_num, column=10, value=bom_count)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
