from fastapi import APIRouter, Depends, HTTPException, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from firebase_admin import firestore
from datetime import datetime
from io import BytesIO
from app.database import get_db
from app import schemas
from app.jinja_templates import templates
from app.firestore_models import document_to_dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()
html_router = APIRouter()


@router.get("/", response_model=list[schemas.Purchase])
def get_purchases(skip: int = 0, limit: int = 100, db = Depends(get_db)):
    purchases_ref = db.collection("purchases")
    docs = purchases_ref.order_by("purchase_date", direction=firestore.Query.DESCENDING).offset(skip).limit(limit).stream()
    purchases = []
    for doc in docs:
        purchase = document_to_dict(doc)
        if purchase:
            purchases.append(purchase)
    return purchases


@router.get("/{purchase_id}", response_model=schemas.Purchase)
def get_purchase(purchase_id: str, db = Depends(get_db)):
    doc = db.collection("purchases").document(purchase_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return document_to_dict(doc)


@router.post("/", response_model=schemas.Purchase)
def create_purchase(purchase: schemas.PurchaseCreate, db = Depends(get_db)):
    # Verify material exists
    material_doc = db.collection("materials").document(purchase.material_id).get()
    if not material_doc.exists:
        raise HTTPException(status_code=404, detail="Material not found")
    
    doc_ref = db.collection("purchases").document()
    data = purchase.model_dump()
    # If qty_remaining not provided, default to qty_purchased
    if data.get("qty_remaining") is None:
        data["qty_remaining"] = data["qty_purchased"]
    data["created_at"] = firestore.SERVER_TIMESTAMP
    doc_ref.set(data)
    doc = doc_ref.get()
    return document_to_dict(doc)


@router.put("/{purchase_id}", response_model=schemas.Purchase)
def update_purchase(purchase_id: str, purchase: schemas.PurchaseUpdate, db = Depends(get_db)):
    doc_ref = db.collection("purchases").document(purchase_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    if purchase.material_id:
        material_doc = db.collection("materials").document(purchase.material_id).get()
        if not material_doc.exists:
            raise HTTPException(status_code=404, detail="Material not found")
    
    update_data = purchase.model_dump(exclude_unset=True)
    if update_data:
        doc_ref.update(update_data)
    
    updated_doc = doc_ref.get()
    return document_to_dict(updated_doc)


@router.delete("/{purchase_id}")
def delete_purchase(purchase_id: str, db = Depends(get_db)):
    doc_ref = db.collection("purchases").document(purchase_id)
    doc = doc_ref.get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Check if any BOM lines reference this purchase
    bom_lines = db.collection("product_bom").where("purchase_id", "==", purchase_id).stream()
    if any(True for _ in bom_lines):
        raise HTTPException(status_code=400, detail="Cannot delete purchase: it is referenced by BOM lines")
    
    doc_ref.delete()
    return {"message": "Purchase deleted"}


# HTML routes
@html_router.get("/purchases", response_class=HTMLResponse)
async def purchases_page(request: Request, page: int = 1, per_page: int = 10, db = Depends(get_db)):
    # Calculate skip
    skip = (page - 1) * per_page
    
    # Get total count
    purchases_ref = db.collection("purchases")
    total_docs = purchases_ref.stream()
    total_count = sum(1 for _ in total_docs)
    
    # Get paginated purchases
    docs = purchases_ref.order_by("purchase_date", direction=firestore.Query.DESCENDING).offset(skip).limit(per_page).stream()
    purchases = []
    for doc in docs:
        purchase = document_to_dict(doc)
        if purchase:
            # Get material name
            material_doc = db.collection("materials").document(purchase["material_id"]).get()
            if material_doc.exists:
                material = document_to_dict(material_doc)
                purchase["material_name"] = material["name"] if material else "Unknown"
            else:
                purchase["material_name"] = "Unknown"
            purchases.append(purchase)
    
    # Calculate pagination info
    total_pages = (total_count + per_page - 1) // per_page if total_count > 0 else 1
    
    # Generate page numbers for pagination (show current page ± 2, plus first and last)
    page_numbers = []
    if total_pages <= 7:
        # Show all pages if 7 or fewer
        page_numbers = list(range(1, total_pages + 1))
    else:
        # Show first page
        page_numbers.append(1)
        # Show pages around current
        start = max(2, page - 2)
        end = min(total_pages - 1, page + 2)
        if start > 2:
            page_numbers.append(None)  # Ellipsis marker
        page_numbers.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_numbers.append(None)  # Ellipsis marker
        # Show last page
        if total_pages > 1:
            page_numbers.append(total_pages)
    
    return templates.TemplateResponse("purchases.html", {
        "request": request,
        "purchases": purchases,
        "page": page,
        "per_page": per_page,
        "total_count": total_count,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "page_numbers": page_numbers
    })


@html_router.post("/purchases/{purchase_id}/delete", response_class=HTMLResponse)
async def delete_purchase_form(request: Request, purchase_id: str, db = Depends(get_db)):
    delete_purchase(purchase_id, db)
    return RedirectResponse(url="/purchases", status_code=303)


# Define delete-all route BEFORE {purchase_id} routes to ensure proper matching
@html_router.post("/purchases/delete-all", response_class=HTMLResponse)
async def delete_all_purchases_form(request: Request, db = Depends(get_db)):
    """Delete all purchases that are not referenced by BOM lines"""
    try:
        purchases_ref = db.collection("purchases")
        purchases = purchases_ref.stream()
        
        deleted_count = 0
        skipped_count = 0
        errors = []
        
        # Collect all purchase IDs first
        purchase_ids = []
        for purchase_doc in purchases:
            if purchase_doc.exists:
                purchase_ids.append(purchase_doc.id)
        
        # If no purchases, just redirect
        if not purchase_ids:
            return RedirectResponse(url="/purchases", status_code=303)
        
        # Delete each purchase (only if not referenced by BOM lines)
        for purchase_id in purchase_ids:
            try:
                # Check if purchase exists
                doc_ref = db.collection("purchases").document(purchase_id)
                doc = doc_ref.get()
                if not doc.exists:
                    continue  # Skip if already deleted
                
                # Check if any BOM lines reference this purchase
                bom_lines = db.collection("product_bom").where("purchase_id", "==", purchase_id).stream()
                bom_count = sum(1 for _ in bom_lines)
                
                if bom_count > 0:
                    # Skip purchases that are referenced by BOM lines
                    skipped_count += 1
                    continue
                
                # Delete purchase
                doc_ref.delete()
                deleted_count += 1
                
            except Exception as e:
                error_msg = str(e)
                # Don't treat "not found" as an error - purchase might have been deleted already
                if "not found" not in error_msg.lower():
                    errors.append(f"Purchase {purchase_id}: {error_msg}")
                    print(f"Error deleting purchase {purchase_id}: {e}")
                # Continue with other purchases even if one fails
        
        if errors:
            print(f"Errors during bulk delete: {errors}")
        
        print(f"Successfully deleted {deleted_count} purchase(s), skipped {skipped_count} (referenced by BOM lines)")
        
    except Exception as e:
        print(f"Critical error in delete_all_purchases_form: {e}")
        import traceback
        traceback.print_exc()
        # Still redirect even if there's an error
    
    return RedirectResponse(url="/purchases", status_code=303)


@html_router.get("/purchases/new", response_class=HTMLResponse)
async def new_purchase_page(request: Request, db = Depends(get_db)):
    materials_ref = db.collection("materials")
    docs = materials_ref.order_by("name").stream()
    materials = []
    for doc in docs:
        material = document_to_dict(doc)
        if material:
            materials.append(material)
    
    # Get unique supplier names from previous purchases
    purchases_ref = db.collection("purchases")
    supplier_set = set()
    for purchase_doc in purchases_ref.stream():
        purchase = purchase_doc.to_dict()
        if purchase and "supplier_name" in purchase:
            supplier_set.add(purchase["supplier_name"])
    supplier_list = sorted(list(supplier_set))
    
    return templates.TemplateResponse("purchase_form.html", {
        "request": request, 
        "purchase": None, 
        "materials": materials,
        "suppliers": supplier_list
    })


@html_router.get("/purchases/{purchase_id}/edit", response_class=HTMLResponse)
async def edit_purchase_page(request: Request, purchase_id: str, db = Depends(get_db)):
    doc = db.collection("purchases").document(purchase_id).get()
    if not doc.exists:
        raise HTTPException(status_code=404, detail="Purchase not found")
    purchase = document_to_dict(doc)
    
    materials_ref = db.collection("materials")
    docs = materials_ref.order_by("name").stream()
    materials = []
    for mat_doc in docs:
        material = document_to_dict(mat_doc)
        if material:
            materials.append(material)
    
    # Get unique supplier names from previous purchases
    purchases_ref = db.collection("purchases")
    supplier_set = set()
    for purchase_doc in purchases_ref.stream():
        purchase_data = purchase_doc.to_dict()
        if purchase_data and "supplier_name" in purchase_data:
            supplier_set.add(purchase_data["supplier_name"])
    supplier_list = sorted(list(supplier_set))
    
    return templates.TemplateResponse("purchase_form.html", {
        "request": request, 
        "purchase": purchase, 
        "materials": materials,
        "suppliers": supplier_list
    })


@html_router.post("/purchases", response_class=HTMLResponse)
async def create_purchase_form(
    request: Request,
    material_id: str = Form(...),
    supplier_name: str = Form(...),
    purchase_date: str = Form(...),
    qty_purchased: float = Form(...),
    unit_cost: float = Form(...),
    currency: str = Form("USD"),
    notes: str = Form(None),
    db = Depends(get_db)
):
    try:
        purchase_date_obj = datetime.fromisoformat(purchase_date.replace("Z", "+00:00"))
    except:
        purchase_date_obj = datetime.fromisoformat(purchase_date)
    
    # qty_remaining defaults to qty_purchased (handled in create_purchase)
    purchase = schemas.PurchaseCreate(
        material_id=material_id,
        supplier_name=supplier_name,
        purchase_date=purchase_date_obj,
        qty_purchased=qty_purchased,
        qty_remaining=None,  # Will default to qty_purchased in create_purchase
        unit_cost=unit_cost,
        currency=currency,
        notes=notes
    )
    create_purchase(purchase, db)
    return RedirectResponse(url="/purchases", status_code=303)


@html_router.post("/purchases/{purchase_id}", response_class=HTMLResponse)
async def update_purchase_form(
    request: Request,
    purchase_id: str,
    material_id: str = Form(...),
    supplier_name: str = Form(...),
    purchase_date: str = Form(...),
    qty_purchased: float = Form(...),
    qty_remaining: float = Form(None),
    unit_cost: float = Form(...),
    currency: str = Form(...),
    notes: str = Form(None),
    db = Depends(get_db)
):
    try:
        purchase_date_obj = datetime.fromisoformat(purchase_date.replace("Z", "+00:00"))
    except:
        purchase_date_obj = datetime.fromisoformat(purchase_date)
    
    purchase = schemas.PurchaseUpdate(
        material_id=material_id,
        supplier_name=supplier_name,
        purchase_date=purchase_date_obj,
        qty_purchased=qty_purchased,
        qty_remaining=qty_remaining,
        unit_cost=unit_cost,
        currency=currency,
        notes=notes
    )
    update_purchase(purchase_id, purchase, db)
    return RedirectResponse(url="/purchases", status_code=303)


@html_router.get("/purchases/export/excel")
async def export_purchases_excel(db = Depends(get_db)):
    """Export all purchases to Excel file"""
    purchases_ref = db.collection("purchases")
    docs = purchases_ref.order_by("purchase_date", direction=firestore.Query.DESCENDING).stream()
    purchases = []
    for doc in docs:
        purchase = document_to_dict(doc)
        if purchase:
            purchases.append(purchase)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"
    
    # Define header row
    headers = [
        "ID",
        "Material",
        "Material Type",
        "Supplier Name",
        "Purchase Date",
        "Quantity Purchased",
        "Quantity Remaining",
        "Unit",
        "Unit Cost",
        "Currency",
        "Total Cost",
        "Notes"
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, purchase in enumerate(purchases, 2):
        # Get material info
        material_doc = db.collection("materials").document(purchase["material_id"]).get()
        if material_doc.exists:
            material = document_to_dict(material_doc)
            material_name = material["name"] if material else "Unknown"
            material_type = material.get("type", "Unknown")
            material_unit = material.get("unit", "")
        else:
            material_name = "Unknown"
            material_type = "Unknown"
            material_unit = ""
        
        total_cost = purchase["qty_purchased"] * purchase["unit_cost"]
        
        # Handle datetime
        purchase_date = purchase.get("purchase_date")
        if purchase_date:
            if hasattr(purchase_date, 'strftime'):
                purchase_date_str = purchase_date.strftime('%Y-%m-%d %H:%M:%S')
            else:
                purchase_date_str = str(purchase_date)
        else:
            purchase_date_str = ""
        
        ws.cell(row=row_num, column=1, value=purchase["id"])
        ws.cell(row=row_num, column=2, value=material_name)
        ws.cell(row=row_num, column=3, value=material_type)
        ws.cell(row=row_num, column=4, value=purchase["supplier_name"])
        ws.cell(row=row_num, column=5, value=purchase_date_str)
        ws.cell(row=row_num, column=6, value=purchase["qty_purchased"])
        ws.cell(row=row_num, column=7, value=purchase.get("qty_remaining", purchase["qty_purchased"]))
        ws.cell(row=row_num, column=8, value=material_unit)
        ws.cell(row=row_num, column=9, value=purchase["unit_cost"])
        ws.cell(row=row_num, column=10, value=purchase["currency"])
        ws.cell(row=row_num, column=11, value=total_cost)
        ws.cell(row=row_num, column=12, value=purchase.get("notes", "") or "")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"purchases_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
